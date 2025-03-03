# Python file that runs in the data server.
# The program gets the data from a database and 
# displays the collected data.
import time
from datetime import datetime
import serial.tools.list_ports
import RPi.GPIO as GPIO
import pandas as pd
import os
from openpyxl import load_workbook
import csv
from flask import Flask, render_template, send_file
import datetime
from flask_httpauth import HTTPDigestAuth
from wtforms import Form, SelectField, StringField, SubmitField, validators
from flask import request, jsonify
from flask_wtf import FlaskForm
from wtforms import StringField, SelectField, SubmitField, validators
from flask import flash, redirect
from flask_socketio import SocketIO, emit
from threading import Lock
from flask import session
from threading import Event
from wtforms import BooleanField
from wtforms.validators import DataRequired
import sys
import paho.mqtt.client as paho
import mysql.connector
from bokeh.plotting import figure, show
from bokeh.embed import components
from bokeh.models.sources import AjaxDataSource
from bokeh.resources import INLINE
from random import randint
from threading import Thread
import simplejson as json
from jinja2 import Template

# Initialize the MQTT client
client = paho.Client()

# Define the "async_mode" variable
async_mode = None

# Initialize the Flask application
app = Flask(__name__)
app.config["CACHE_TYPE"] = "null"
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0

# Initialize socketio with Flask app
socketio = SocketIO(app, async_mode=async_mode, transports=["polling"])

# Initialize thread variables
thread = None
thread_lock = Lock()
thread_event = Event()

# Set the secret key for session management
app.config['SECRET_KEY'] = 'secret key here'

# Initialize HTTP Digest Authentication
auth = HTTPDigestAuth()

# Initialize serial communication
ser = serial.Serial()

# Global variable for storing IP address
ip_nariz=""

# User credentials for authentication
users = {
    "nariz": "electronica"
}

# Global variable for storing the database name
basededatos=""

# Form for selecting a database and performing actions on it
class SimpleForm2(FlaskForm):
    analisis = SelectField('Seleciona base de datos', choices=[], coerce=int)
    accion = SelectField('Acción sobre base de datos', choices=[(1, 'Acceder'), (2, 'Borrar')])
    submit = SubmitField('Submit')
    
    def __init__(self, accounts=None):
        super().__init__()  # calls the base initialization
        if accounts:
            self.analisis.choices = [(i, c) for i, c in enumerate(accounts)]

# Form for selecting a table and performing actions on it
class SimpleForm(FlaskForm):
    analisis = SelectField('Seleciona tabla', choices=[], coerce=int)
    accion = SelectField('Acción sobre tabla', choices=[(1, 'Visualizar'), (2, 'Borrar'), (3, 'Descargar')])
    submit = SubmitField('Submit')
    
    def __init__(self, accounts=None):
        super().__init__()  # calls the base initialization
        if accounts:
            self.analisis.choices = [(i, c) for i, c in enumerate(accounts)]

# Function for getting the password for authentication
@auth.get_password
def get_pw(username):
    if username in users:
        return users.get(username)
    return None

# Route for the home page
@app.route("/", methods=['GET', 'POST'])
@auth.login_required
def hello():
    global Tabla_selected, id_ip_nariz, basededatos, starting_time
    stmt = "SHOW DATABASES"
    mydb = mysql.connector.connect(
        host="localhost",
        user="rasbe",
        password="mas12%i&&",
    )
    
    dts = round((datetime.datetime.now() - starting_time).total_seconds())
    
    nuevaiteracion = True
    
    while(nuevaiteracion):
        nuevaiteracion = False
        ip_nariz=""
        if (len(id_ip_nariz) == 0):
            starting_time = datetime.datetime.now()
        else:
            for id, ip in id_ip_nariz.items():
                if ip[1] > (dts - 60):
                    ip_nariz += id + " : " + ip[0]
                else:
                    nuevaiteracion = True
                    del id_ip_nariz[id]
                    break
                
    print("")
    print("Id_ip_nariz")
    print(id_ip_nariz)
    print("ip_nariz:")
    print(ip_nariz)
 
    mycursor = mydb.cursor()
    print(stmt)
    mycursor.execute(stmt)
    result = mycursor.fetchall()
    if not result:
        return redirect('/sindatos')
    
    result2 = []
    for i in result:
        result2.append(str(i))
    result2 = [e.replace("(", "").replace(")", "").replace("'", "").replace(",", "") if isinstance(e, str) and e[0] == '(' else e  for e in result2]

    print("result2:")
    print(result2)
    result2.remove('information_schema')
    result2.remove('mysql')
    result2.remove('performance_schema')

    user_form = SimpleForm2(accounts=result2)

    templateData = {
      'title': 'Gestión Base de Datos',
      'elementos': len(result2),
      'base_de_datos': result2,
      'ip_nariz': ip_nariz
    }      
    
    if user_form.validate_on_submit():
        print("Análisis: {}".format(user_form.analisis.data))
        print("Acción: {}".format(user_form.accion.data))
        Tabla = user_form.analisis.data
        Accion = user_form.accion.data
        if Accion == '1':
            basededatos = result2[Tabla]
            print("Basededatos:" + basededatos)
            return redirect('/tablas')
        if Accion == '2':
            stmt = "DROP DATABASE " + result2[Tabla]
            print(stmt)
            mycursor.execute(stmt)
            mydb.commit()
            return redirect('/')
        else:
            return redirect('/')
    mydb.close()
    return render_template('index.html', **templateData, form=user_form)

# Route for displaying tables in the selected database
@app.route("/tablas", methods=['GET', 'POST'])
def tablas():
    global Tabla_selected, ip_nariz, basededatos
    stmt = "SHOW TABLES"
    mydb = mysql.connector.connect(
        host="localhost",
        user="rasbe",
        password="mas12%i&&",
        database=basededatos
    )
    mycursor = mydb.cursor()
    print(stmt)
    mycursor.execute(stmt)
    result = mycursor.fetchall()
    if not result:
        return redirect('/sindatos')
    
    result2=[]
    for i in result:
        result2.append(str(i))
    result2 = [e.replace("(", "").replace(")", "").replace("'", "").replace(",", "") if isinstance(e, str) and e[0] == '(' else e  for e in result2]

    print("result2:")
    print(result2)

    user_form = SimpleForm(accounts=result2)

    templateData = {
        'title': 'Gestión de las tablas',
        'elementos': len(result2),
        'tablas': result2,
        'ip_nariz': ip_nariz
    }

    if user_form.validate_on_submit():
        print("Análisis: {}".format(user_form.analisis.data))
        print("Acción: {}".format(user_form.accion.data))
        Tabla = user_form.analisis.data
        Accion = user_form.accion.data
        if Accion == '1':
            Tabla_selected = result2[Tabla]
            return redirect('/dashboard')
        if Accion == '2':
            stmt = "DROP TABLE " + result2[Tabla]
            print(stmt)
            mycursor.execute(stmt)
            mydb.commit()
            return redirect('/')
        if Accion == '3':
            Tabla_selected = result2[Tabla]
            return redirect('/descargar')
        else:
            return redirect('/')
    mydb.close()
    return render_template('tablas.html', **templateData, form=user_form)

# Route for displaying a message when there is no data
@app.route("/sindatos", methods=['GET'])
def sindatos():
    global Tabla_selected, ip_nariz
    templateData = {
        'title': 'Base de Datos Vacía',
        'ip_nariz': ip_nariz
    }
    return render_template('sindatos.html', **templateData)

# Route for downloading data from the selected table
@app.route("/descargar", methods=['GET'])
def descargar():
    global Tabla_selected, basededatos
    # Specify the file name
    file_name_xls = "/home/rasbe/Documentos/Python/TABLEMULTIPLOT/" + Tabla_selected + ".xlsx"
    
    dir = "/home/rasbe/Documentos/Python/TABLEMULTIPLOT/"

    # Remove existing .xlsx files in the directory
    lista_ficheros = os.listdir(dir)
    for fichero in lista_ficheros:
        if fichero.endswith(".xlsx"):
            file_name_remove = "/home/rasbe/Documentos/Python/TABLEMULTIPLOT/" + fichero
            print(fichero)
            print(file_name_remove)
            os.remove(file_name_remove)

    column_names = ['Sample time', 'Power', 'Mq135', 'Mq2', 'Mq3', 'Mq4', 'Mq5', 'Mq9', 'Mq7', 'Mq8', 'TempDS', 'TempDHT', 'Humidity', 'Sample Name']

    # Create a DataFrame from the list
    df = pd.DataFrame(column_names).T
    df.to_excel(file_name_xls, index=False, header=False)
    
    mydb = mysql.connector.connect(
        host="localhost",
        user="rasbe",
        password="mas12%i&&",
        database=basededatos
    )
 
    mycursor = mydb.cursor()
    print("Tabla_selected")
    print(Tabla_selected)
    stmt = "SELECT * FROM " + Tabla_selected + " ORDER BY Ind"
    print(stmt)
    mycursor.execute(stmt)
    result = mycursor.fetchall()
    
    book = load_workbook(file_name_xls)
    # Get the last row in the existing Excel sheet
    # If it doesn't exist, append a header row
    if 'Sheet1' in book.sheetnames:
            sheet = book['Sheet1']

    for i in result:
        str_list = ['' for i in range(1, 15)]
        # Convert and store the elements of arr to a list of strings
        str_list[0] = str(i[1])
        str_list[1] = str(i[2])
        str_list[2] = str(i[3])
        str_list[3] = str(i[4])
        str_list[4] = str(i[5])
        str_list[5] = str(i[6])
        str_list[6] = str(i[7])
        str_list[7] = str(i[8])
        str_list[8] = str(i[9])
        str_list[9] = str(i[10])
        str_list[10] = str(i[11])
        str_list[11] = str(i[12])
        str_list[12] = str(i[13])
        str_list[13] = str(i[14])

        # Append list to the sheet
        sheet.append(str_list)
        
    book.save(file_name_xls)
          
    mydb.close()
    return send_file(file_name_xls, as_attachment=True)

# Global variable for storing the selected table name
Tabla_selected=""

# Template for displaying sensor data
template = Template('''<!DOCTYPE html>
<html lang="en">
    <head>
        <meta charset="utf-8">
        <title>Datos de los sensores</title>
        <meta name="viewport" content="width=device-width, initial-scale=1">
        <link rel="stylesheet" href="https://maxcdn.bootstrapcdn.com/bootstrap/3.4.1/css/bootstrap.min.css">
        {{ js_resources }}
    </head>
    <body>
    {{ plot_div }}
    {{ plot_script }}
    <h2 class="m-5"> 
        Muestreo de la nariz:  {{Nariz_ID}}
    </h2>
    <h2 class="m-5">
        Fecha del muestreo: {{Fecha}}
    </h2> 
    <h2 class="m-5">
        <input type="button" value="Volver" onclick="location.href='/tablas'">
    </h2> 
    </body>
</html>
''')

# Route for displaying the dashboard with sensor data
@app.route("/dashboard", methods=['GET'])
def dashboard():
    mydb = mysql.connector.connect(
        host="localhost",
        user="rasbe",
        password="mas12%i&&",
        database=basededatos
    )
 
    mycursor = mydb.cursor()
    print("Tabla_selected")
    print(Tabla_selected)
    stmt = "SELECT Nariz_ID, Fecha FROM " + Tabla_selected + " WHERE Ind = 1"
    print(stmt)
    mycursor.execute(stmt)
    result = mycursor.fetchall()
    print("result:")
    print(result)
    Nariz_ID = result[0][0]
    Fecha = result[0][1].replace("T", " / ")
    
    streaming = True
    source1 = AjaxDataSource(data_url=request.url_root + "/data1", polling_interval=2000, mode='replace')
    source1.data = dict(x=[], y1=[])

    source2 = AjaxDataSource(data_url=request.url_root + "/data2", polling_interval=2000, mode='replace')
    source2.data = dict(x=[], y2=[])

    source3 = AjaxDataSource(data_url=request.url_root + "/data3", polling_interval=2000, mode='replace')
    source3.data = dict(x=[], y3=[])

    source4 = AjaxDataSource(data_url=request.url_root + "/data4", polling_interval=2000, mode='replace')
    source4.data = dict(x=[], y4=[])

    source5 = AjaxDataSource(data_url=request.url_root + "/data5", polling_interval=2000, mode='replace')
    source5.data = dict(x=[], y5=[])

    source6 = AjaxDataSource(data_url=request.url_root + "/data6", polling_interval=2000, mode='replace')
    source6.data = dict(x=[], y6=[])

    source7 = AjaxDataSource(data_url=request.url_root + "/data7", polling_interval=2000, mode='replace')
    source7.data = dict(x=[], y7=[])

    source8 = AjaxDataSource(data_url=request.url_root + "/data8", polling_interval=2000, mode='replace')
    source8.data = dict(x=[], y8=[])

    fig = figure(title="Datos de los sensores", x_axis_label='Muestras', y_axis_label="Valor", height=300, sizing_mode='scale_width')
    fig.line('x', 'y1', source=source1, legend_label="Mq135", line_color="blue", line_width=2)
    fig.line('x', 'y2', source=source2, legend_label="Mq2", line_color="red", line_width=2)
    fig.line('x', 'y3', source=source3, legend_label="Mq3", line_color="brown", line_width=2)
    fig.line('x', 'y4', source=source4, legend_label="Mq4", line_color="green", line_width=2)
    fig.line('x', 'y5', source=source5, legend_label="Mq5", line_color="black", line_width=2)
    fig.line('x', 'y6', source=source6, legend_label="Mq9", line_color="cyan", line_width=2)
    fig.line('x', 'y7', source=source7, legend_label="Mq7", line_color="pink", line_width=2)
    fig.line('x', 'y8', source=source8, legend_label="Mq8", line_color="grey", line_width=2)

    js_resources = INLINE.render_js()
    css_resources = INLINE.render_css()

    script, div = components(fig, INLINE)

    html = template.render(
        plot_script=script,
        plot_div=div,
        js_resources=js_resources,
        css_resources=css_resources,
        Nariz_ID=Nariz_ID,
        Fecha=Fecha
    )
    mydb.close()
    return html

result=[]

# Route for fetching data for sensor Mq135
@app.route('/data1', methods=['GET', 'POST'])
def data1():
    global Tabla_selected, result, basededatos
    mydb = mysql.connector.connect(
        host="localhost",
        user="rasbe",
        password="mas12%i&&",
        database=basededatos
    )
 
    mycursor = mydb.cursor()
    print("Tabla_selected")
    print(Tabla_selected)
    stmt = "SELECT * FROM " + Tabla_selected + " ORDER BY Ind DESC LIMIT 200"
    print(stmt)
    mycursor.execute(stmt)
    result = mycursor.fetchall()
    Mq135 = []
    Sample = []
    for i in result:
        Mq135.append(i[3])
        Sample.append(i[1])
    mydb.close()
    return jsonify(x=Sample, y1=Mq135)

# Route for fetching data for sensor Mq2
@app.route('/data2', methods=['GET', 'POST'])
def data2():
    global result 
    Mq2 = []
    Sample = []
    for i in result:
        Mq2.append(i[4])
        Sample.append(i[1])
    return jsonify(x=Sample, y2=Mq2)

# Route for fetching data for sensor Mq3
@app.route('/data3', methods=['GET', 'POST'])
def data3():
    global result 
    Mq3 = []
    Sample = []
    for i in result:
        Mq3.append(i[5])
        Sample.append(i[1])
    return jsonify(x=Sample, y3=Mq3)

# Route for fetching data for sensor Mq4
@app.route('/data4', methods=['GET', 'POST'])
def data4():
    global result 
    Mq4 = []
    Sample = []
    for i in result:
        Mq4.append(i[6])
        Sample.append(i[1])
    return jsonify(x=Sample, y4=Mq4)

# Route for fetching data for sensor Mq5
@app.route('/data5', methods=['GET', 'POST'])
def data5():
    global result 
    Mq5 = []
    Sample = []
    for i in result:
        Mq5.append(i[7])
        Sample.append(i[1])
    return jsonify(x=Sample, y5=Mq5)

@app.route('/data6', methods=['GET','POST'])
def data6():
    global result 
    Mq9 = []
    Sample = []
    for i in result:
        Mq9.append(i[8])
        Sample.append(i[1])  
    # print("Mq9:")
    # print(Mq9)
    # print("Sample:")
    # print(Sample)
    return jsonify(x=Sample,y6=Mq9)

@app.route('/data7', methods=['GET','POST'])
def data7():
    global result 
    Mq7 = []
    Sample = []
    for i in result:
        Mq7.append(i[9])
        Sample.append(i[1])  
    # print("Mq7:")
    # print(Mq7)
    # print("Sample:")
    # print(Sample)
    return jsonify(x=Sample,y7=Mq7)

@app.route('/data8', methods=['GET','POST'])
def data8():
    global result 
    Mq8 = []
    Sample = []
    for i in result:
        Mq8.append(i[10])
        Sample.append(i[1])  
    # print("Mq8:")
    # print(Mq8)
    # print("Sample:")
    # print(Sample)
    return jsonify(x=Sample,y8=Mq8)

id_ip_nariz = {}

starting_time = 0

@app.route('/ip', methods=['POST'])
def json():
    global id_ip_nariz, ip_nariz, starting_time
    data = request.get_json()
    print("")
    print("IP recibida")
    print(data)
    Nariz_ID=data.get('Nariz_ID')
    #print("Nariz_ID:")
    #print(Nariz_ID)
    Nariz_IP=data.get('mi_ip')
    #print("IP Nariz:")
    #print(Nariz_IP)
    dt = round((datetime.datetime.now() - starting_time).total_seconds())
    id_ip_nariz[Nariz_ID] = [Nariz_IP, dt]
    print("Id_ip_nariz")
    print(id_ip_nariz)
    print("")
    return jsonify(data)




if __name__ == "__main__":
    
    
    
    starting_time = datetime.datetime.now()
    app.run(host='0.0.0.0', port=8080, debug=True)




