import time
from datetime import datetime
import serial.tools.list_ports
import RPi.GPIO as GPIO
import pandas as pd
import os
from openpyxl import load_workbook
import csv
from flask import Flask, render_template, send_file
import datetime
from flask_httpauth import HTTPDigestAuth
from wtforms import Form, SelectField, StringField, SubmitField, validators
from flask import request
from flask_wtf import FlaskForm
from wtforms import StringField, SelectField, SubmitField, validators
from flask import flash, redirect
from flask_socketio import SocketIO, emit
from threading import Lock, Thread
from flask import session
from threading import Event
from wtforms import BooleanField
from wtforms.validators import DataRequired
import sys
from urllib import request
import paho.mqtt.client as paho
from luma.core.interface.serial import i2c
from luma.core.render import canvas
from luma.oled.device import sh1106
import subprocess
from PIL import Image, ImageDraw, ImageFont
import requests


client = paho.Client()

async_mode = None  # Define the "async_mode" variable

app = Flask(__name__)
app.config["CACHE_TYPE"] = "null"
socketio = SocketIO(app, async_mode=async_mode, transports=["polling"])

thread = None
thread_lock = Lock()
thread_event = Event()
app.config['SECRET_KEY'] = 'secret key here'
auth = HTTPDigestAuth()
ser = serial.Serial()
serialport = None

users = {
    "nariz": "electronica"
}

Nariz_ID = "Nariz1"

seriali2c = i2c(port=1, address=0x3C)

device = sh1106(seriali2c)
device.contrast(255)

font_path = str('/home/rasbe/Documentos/Codigo_python/oled-sh1106/opt/oled/fonts/Roboto-Light.ttf')
font12 = ImageFont.truetype(font_path, 11)
font14 = ImageFont.truetype(font_path, 14)
font20 = ImageFont.truetype(font_path, 20)
font = ImageFont.truetype(font_path, 18)

class SimpleForm(FlaskForm):
    nombrefichero = StringField('Nombre del fichero Excel ', [validators.Length(min=1, max=100)])
    nombremuestras = StringField('Nombre de las muestras ', [validators.Length(min=1, max=100)])
    check = BooleanField('Subir datos al servidor ') #, validators=[DataRequired()])
    basedatos = StringField('Nombre de la base de datos ', [validators.Length(min=1, max=100)])
    tabla = StringField('Nombre de la tabla en la base de datos ', [validators.Length(min=1, max=100)])
    ipservidor = StringField('IP Servidor', [validators.Length(min=1, max=100)])
    analisis = SelectField('Analisis', choices=[(1, 'Real Time'), (2, 'ADC'), (3, 'Calibration'), (4, 'eNose Study')])
    submit = SubmitField('Submit')

def calcTime(enter,exit):
    format="%H:%M:%S"
    #Parsing the time to str and taking only the hour,minute,second (without miliseconds)
    enterStr = str(enter).split(".")[0]
    exitStr = str(exit).split(".")[0]
    #Creating enter and exit time objects from str in the format (H:M:S)
    enterTime = datetime.strptime(enterStr, format)
    exitTime = datetime.strptime(exitStr, format)
    return exitTime - enterTime

def get_ip():
    global Nariz_ID
    cmd = "hostname -I | awk '{print $2}'"
    IP = subprocess.check_output(cmd, shell = True ).decode("utf-8")
    url='http://192.168.0.251:8080/ip'
    myobj = {'Nariz_ID': Nariz_ID, 'mi_ip':IP}
    try:
        print(myobj)
        x = requests.post(url,json=myobj, timeout=5)
        print(x.text)
        return "IP:  " + str(IP).rstrip("\r\n")+"\r\nCon conexion al servidor\r\n"
    except:
        print("No se pudo enviar IP")
        return "IP:  " + str(IP).rstrip("\r\n")+"\r\nSin conexion al servidor\r\n"

def internet_connection():
    try:
        return subprocess.run(['wget','-q','--spider','www.goole.com'],timeout=5).returncode == 0
    except subprocess.TimeoutExpired:
        return False

def reboot():
    try:
        return os.system('sudo reboot')
    except:
        print("No se pudo reiniciar")
        return False

def rebooting():
    while True:
        reintentos = 0
        while internet_connection()==False:
            msg = str("Sin conexion a Internet")
            print("Sin conexion a Internet")
            msg2 = get_ip()
            with canvas(device) as draw:
                draw.text((0, 0),  msg, font=font12, fill=255)
                draw.text((0, 15),  msg2, font=font12, fill=255)
            time.sleep(5)
            reintentos = reintentos + 1
            if reintentos == 5:
                print("reboot")
                reboot()

        with canvas(device) as draw:
                    msg = "Con conexion a Internet"
                    print("Con conexion a Internet")
                    msg2 = get_ip()
                    msg3 = "Servidor local preparado"
                    draw.text((0, 0),  msg, font=font12, fill=255)
                    draw.text((0, 15),  msg2, font=font12, fill=255)
                    draw.text((0, 45),  msg3, font=font12, fill=255)
        time.sleep(20)
          


@auth.get_password
def get_pw(username):
    if username in users:
        return users.get(username)
    return None

@app.route("/")
@auth.login_required
def hello():
    global thread, Nariz_ID
    now = datetime.datetime.now()
    timeString = now.strftime("%Y-%m-%d %H:%M")
    templateData = {
      'title' : 'Nariz Electrónica',
      'time': timeString,
      'Nariz_ID': Nariz_ID
      }

    print('Desconectar si socket conectado')
    thread_event.clear()
    with thread_lock:
        if thread is not None:
            thread.join()
            thread = None

    return render_template('index.html', **templateData)

@app.route("/conectar_nariz", methods=['GET', 'POST'])
def conectar_nariz():
    # Use the Broadcom SOC channel, which maps pin numbers like on the Pi
    global serialport

    GPIO.setwarnings(False)
    GPIO.setmode(GPIO.BCM)

    # Set up GPIO26 as an output
    GPIO.setup(26, GPIO.OUT)

    # Set GPIO26 to low
    GPIO.output(26, GPIO.LOW)
    time.sleep(1)  # wait for 1 second

    # Set GPIO26 to high
    GPIO.setup(26, GPIO.IN) 
    time.sleep(10)  # wait for 10 second

    ports = serial.tools.list_ports.comports()

    for port, desc, hwid in sorted(ports):
        print("{}: {}".format(port, desc))
        time.sleep(1)  # delay for 1 secondelay

    arduino_port = None

    for port, desc, hwid in sorted(ports):
        if 'USB2.0-Serial' in desc:
            arduino_port = port
            serialport = port
            break
            
    if arduino_port:
        print("Nariz encontrada en: {}".format(arduino_port))
        conectaString = "Nariz encontrada en {}".format(arduino_port)
        templateData = {
            'title' : 'Conexión',
            'conecta': conectaString,
            'siguiente': 1,
        }

    else:
        print("Nariz no encontrada")
        conectaString = "Nariz no encontrada"
        templateData = {
            'title' : 'Conexión',
            'conecta': conectaString,
            'siguiente': 0,
        }
    
    return render_template('conectado.html', **templateData)

Nombre_fichero = ""
Nombre_muestras = ""
Tipo_analisis = 1
Ip_Servidor = ""
Enviar_nube = False
Tabla = ""
Basedatos = ""

@app.route("/form", methods=['GET', 'POST'])
def form():
    global Nombre_fichero, Nombre_muestras, Tipo_analisis, Ip_Servidor, Enviar_nube, Tabla, Basedatos
    user_form = SimpleForm()
    templateData = {
            'title' : 'Formulario',
    }

    if user_form.validate_on_submit():
        print("Nombre fichero: " + user_form.nombrefichero.data)
        print("Nombre muestras: " + user_form.nombremuestras.data)
        print("Tipo análisis: {}".format(user_form.analisis.data))
        print("Check envío al servidor: {}".format(user_form.check.data))
        print("Base de datos: {}".format(user_form.basedatos.data))
        print("Tabla: {}".format(user_form.tabla.data))
        print("IP Servidor: {}".format(user_form.ipservidor.data))
        Nombre_fichero = user_form.nombrefichero.data
        Nombre_muestras = user_form.nombremuestras.data
        Tipo_analisis = user_form.analisis.data
        Enviar_nube = user_form.check.data
        Tabla = user_form.tabla.data
        Basedatos = user_form.basedatos.data
        Ip_Servidor = user_form.ipservidor.data
        return redirect('/sockets')

    return render_template('form.html', **templateData, form=user_form)

def background_thread(event):
    """Example of how to send server generated events to clients."""
    global thread
    global Nariz_ID, Tabla, Basedatos
    global Nombre_fichero, Nombre_muestras, Tipo_analisis, Enviar_nube, Ip_Servidor, ser
    count = 1
    file = 0
    ser = 0
    host = Ip_Servidor

    if Enviar_nube == True:
        client.username_pw_set("nariz", "electronica")
        try:
            client.connect(host, 1883, 60)

        except:
            print("Couldn't connect to the mqtt broker")
            socketio.emit('error_message', {'data': 'Error de conexión con broker MQTT'})
            return

    print("Connected to the mqtt broker")

    muestreo_ini()

    # Obtener la hora actual
    enter = datetime.datetime.now()

    # Imprimir la hora actual
    print("La hora actual es:", enter)
    
    print("Puerto serie")
    print(serialport)
    
    ser=serial.Serial(serialport,115200)
    ser.setDTR(0)
    time.sleep(1)  # wait for 1 second
    ser.setDTR(1)
    
    # Set up GPIO26 as an output
    GPIO.setup(26, GPIO.OUT)

    # Set GPIO26 to low
    GPIO.output(26, GPIO.LOW)
    time.sleep(1)  # wait for 1 second

    # Set GPIO26 to high
    GPIO.setup(26, GPIO.IN)
    time.sleep(10)  # wait for 1 second

    Tipo = int(Tipo_analisis)
    print(Tipo)
    print(type(Tipo))
    a = Tipo.to_bytes(1 ,'big')
    print(a)
    print(ser)
    ser.timeout = 5
    ser.reset_input_buffer()  # Flush the input buffer
    ser.write(a)
    ser.flush()
    
    file_name = "/home/rasbe/Documentos/Codigo_python/USB/"+Nombre_fichero+".csv"
    file_name_xls = "/home/rasbe/Documentos/Codigo_python/USB/"+Nombre_fichero+".xlsx"

    arr = bytearray([1]*27)
    i = 0

    try:
        while event.is_set():
            while i<25 and event.is_set():

                readedText = ser.read()
        
                print(" ", end='')
                print(hex(int.from_bytes(readedText, byteorder='big')), end='')

                arr[i] = int.from_bytes(readedText, byteorder='big')

                i = i+1

                if i == 25:
                    arr[i] = 0
                    i+=1
                    arr[i] = 0
                    print("")
                    str_list = [int(b) for b in arr]
                    str_list[26] = 'FA'

                    # Append list to the CSV file
                    file = open(file_name, 'a', newline='')
                    writer = csv.writer(file)
                    writer.writerow(str_list)
                    file.close()

                    row = str_list

                    book = load_workbook(file_name_xls)
                    # # Get the last row in the existing Excel sheet
                    # # If it doesn't exist, append a header row
                    if 'Sheet1' in book.sheetnames:
                        sheet = book['Sheet1']

                    str_list = ['' for i in range(1, 15)]
                    # Convert and store the elements of arr to a list of strings
                    str_list[0] = str(int(row[1])*256 + int(row[0]))
                    str_list[1] = row[2]
                    str_list[2] = str(int(row[4])*256 + int(row[3]))
                    str_list[3] = str(int(row[6])*256 + int(row[5]))
                    str_list[4] = str(int(row[8])*256 + int(row[7]))
                    str_list[5] = str(int(row[10])*256 + int(row[9]))
                    str_list[6] = str(int(row[12])*256 + int(row[11]))
                    str_list[7] = str(int(row[14])*256 + int(row[13]))
                    str_list[8] = str(int(row[16])*256 + int(row[15]))
                    str_list[9] = str(int(row[18])*256 + int(row[17]))
                    str_list[10] = str(int(row[20])/10 + int(row[19]))
                    str_list[11] = str(int(row[22])/10 + int(row[21]))
                    str_list[12] = str(int(row[24])/10 + int(row[23]))
                    str_list[13] = row[26]

                    # Append list to the sheet
                    sheet.append(str_list)
                    book.save(file_name_xls)
                    mensaje_json = '{"Index": '+ str(count) +', "Sample time": ' + str(int(row[1])*256 + int(row[0])) + ', "Power": ' + str(row[2]) + ', "Mq135": ' + str(int(row[4])*256 + int(row[3])) + ', "Mq2": ' + str(int(row[6])*256 + int(row[5])) + ', "Mq3": ' + str(int(row[8])*256 + int(row[7])) + ', "Mq4": ' + str(int(row[10])*256 + int(row[9])) + ', "Mq5": ' + str(int(row[12])*256 + int(row[11])) + ', "Mq9": ' + str(int(row[14])*256 + int(row[13])) + ', "Mq7": ' + str(int(row[16])*256 + int(row[15])) + ', "Mq8": ' + str(int(row[18])*256 + int(row[17])) + ', "TempDS": ' + str(int(row[20])/10 + int(row[19])) + ', "TempDHT": ' + str(int(row[22])/10 + int(row[21])) + ', "Humidity": ' + str(int(row[24])/10 + int(row[23])) + ', "Sample Name": "' + row[26] + '"}'
                    Topic = Nariz_ID + '/' + Basedatos + '/' + Tabla
                    print("Topic:" + Topic)
                    client.publish(Topic, mensaje_json, 0)

            i = 0

            print('Background emit ' + str(count))
            exit = datetime.datetime.now()
            duration = datetime.timedelta(hours=exit.hour-enter.hour, minutes=exit.minute-enter.minute, seconds=exit.second-enter.second)
            print("Tiempo transcurrido:" + str(duration))
            socketio.emit('my_response', {'data': str(duration) , 'count': count})
            count += 1
            with canvas(device) as draw:
                msg = "Enviando datos"
                msg2 = str(count)
                draw.text((0, 0),  msg, font=font12, fill=255)
                draw.text((0, 15),  msg2, font=font12, fill=255)

    finally:
        event.clear()
        thread = None
        with canvas(device) as draw:
                msg = "Con conexion a Internet"
                print("Con conexion a Internet")
                msg2 = get_ip()
                msg3 = "Servidor preparado"
                draw.text((0, 0),  msg, font=font12, fill=255)
                draw.text((0, 15),  msg2, font=font12, fill=255)
                draw.text((0, 30),  msg3, font=font12, fill=255)
        print('Background_exit ' + str(count))
        if Enviar_nube == True:
            client.disconnect()
        if file:
            file.close()
        if ser:
            ser.close()

@app.route("/sockets")
def sockets():
    return render_template('sockets.html', async_mode=socketio.async_mode)

#@socketio.event
@socketio.on('my_event')
def my_event(message):
    print('Event '+ message['data'])
    session['receive_count'] = session.get('receive_count', 0) + 1
    duration = datetime.timedelta(hours=0, minutes=0, seconds=0)
    emit('my_response',  {'data': str(duration), 'count': 0})

# Receive the test request from client and send back a test response
@socketio.on('stop_message')
def handle_message(data):
    print('received message: ' + str(data))
    emit('stop_message', {'data': 'Se ha enviado la señal de parada'})

@socketio.on('disconnect')
def test_disconnect():
    print('Client disconnected')
    global thread
    thread_event.clear()
    with thread_lock:
        if thread is not None:
            thread.join()
            thread = None

#@socketio.event
@socketio.on('connect')
def connect():
    global thread
    print('Connected first time')
    with thread_lock:
        if thread is None:
            thread_event.set()
            thread = socketio.start_background_task(background_thread, thread_event)

def muestreo_ini():
    global Nombre_fichero, Nombre_muestras, Tipo_analisis, ser

    # Specify the file name
    file_name = "/home/rasbe/Documentos/Codigo_python/USB/"+Nombre_fichero+".csv"
    file_name_xls = "/home/rasbe/Documentos/Codigo_python/USB/"+Nombre_fichero+".xlsx"

    dir = "/home/rasbe/Documentos/Codigo_python/USB/"
    lista_ficheros = os.listdir(dir)
    for fichero in lista_ficheros:
        if fichero.endswith(".csv"):
            file_name_remove = "/home/rasbe/Documentos/Codigo_python/USB/" + fichero
            print(fichero)
            print(file_name_remove)
            os.remove(file_name_remove)
        if fichero.endswith(".xlsx"):
            file_name_remove = "/home/rasbe/Documentos/Codigo_python/USB/" + fichero
            print(fichero)
            print(file_name_remove)
            os.remove(file_name_remove)

    # Create a list of strings
    column_names = ['C_' + str(i) for i in range(1, 28)]
    
    
    file = open( file_name, 'w', newline='')
    writer = csv.writer(file)
    writer.writerow(column_names)
    file.close()

    column_names = ['Sample time', 'Power', 'Mq135', 'Mq2', 'Mq3', 'Mq4', 'Mq5', 'Mq9', 'Mq7', 'Mq8', 'TempDS', 'TempDHT', 'Humidity', 'Sample Name']

    # Create a DataFrame from the list
    df = pd.DataFrame(column_names).T
    df.to_excel(file_name_xls, index=False, header=False)

@app.route("/completado")
def completado():

    templateData = {
      'title' : 'Muestreo completado',
    } 
    
    ser.setDTR(0)
    time.sleep(1)  # wait for 1 second
    ser.setDTR(1)
    # Set up GPIO26 as an output
    # GPIO.setup(26, GPIO.OUT)

    # Set GPIO26 to low
    # GPIO.output(26, GPIO.LOW)
    # time.sleep(1)  # wait for 1 second

    # Set GPIO26 to high
    # GPIO.setup(26, GPIO.IN) 
    time.sleep(10)  # wait for 10 second

    return render_template('completado.html', **templateData)

@app.route("/descargar")
def descargar():
    global Nombre_fichero
    path = "/home/rasbe/Documentos/Codigo_python/USB/"+Nombre_fichero+".xlsx"
    return send_file(path, as_attachment=True)


if __name__ == "__main__":

    thread1 = socketio.start_background_task(rebooting)
    socketio.run(app, host='0.0.0.0', port=8080, allow_unsafe_werkzeug=True) #, debug=True)




